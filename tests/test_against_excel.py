"""Verify Python model matches cached values from sprint_planning.xlsx."""

from __future__ import annotations

import math
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sprint_forecast.excel_loader import load_excel_params, load_plan_from_excel
from sprint_forecast.model import compute_forecast

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "sprint_planning.xlsx"


def _close(a, b, rel=1e-6, abs_tol=1e-6):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, float) and math.isnan(a):
        return b is None or (isinstance(b, float) and math.isnan(b))
    return math.isclose(float(a), float(b), rel_tol=rel, abs_tol=abs_tol)


@pytest.fixture(scope="module")
def forecast():
    items = load_plan_from_excel(EXCEL)
    params = load_excel_params(EXCEL)
    return compute_forecast(
        items,
        first_sprint_start=params["first_sprint_start"],
        sprint_length_days=params["sprint_length_days"],
        sprint_count=15,
    )


@pytest.fixture(scope="module")
def excel_formulas():
    return load_workbook(EXCEL, data_only=True)["Formulas"]


def test_summary_matches_excel(forecast, excel_formulas):
    ws = excel_formulas
    assert forecast.initial_backlog_size == ws["B1"].value
    assert forecast.backlog_size == ws["B2"].value
    assert forecast.sprint_length_days == int(ws["B3"].value)
    report = ws["B4"].value
    if isinstance(report, datetime):
        report = report.date()
    assert forecast.report_date == report
    assert forecast.current_sprint == ws["B5"].value
    assert forecast.remaining_work == ws["B6"].value
    assert _close(forecast.velocity_mean, ws["B8"].value)
    assert _close(forecast.velocity_std, ws["C8"].value, abs_tol=1e-6)
    assert _close(forecast.addlocity_mean, ws["B9"].value)
    assert _close(forecast.addlocity_std, ws["C9"].value, abs_tol=1e-6)


def test_sprint_rows_match_excel(forecast, excel_formulas):
    ws = excel_formulas
    for sprint in forecast.sprints:
        row = 12 + sprint.sprint_number
        assert sprint.sprint_number == int(ws[f"A{row}"].value)

        if sprint.sprint_number >= 1:
            start = ws[f"B{row}"].value
            end = ws[f"C{row}"].value
            if isinstance(start, datetime):
                start = start.date()
            if isinstance(end, datetime):
                end = end.date()
            assert sprint.start == start
            assert sprint.end == end

        assert _close(sprint.done, ws[f"D{row}"].value)

        excel_total_done = ws[f"E{row}"].value
        if excel_total_done in ("", None):
            assert sprint.total_added is not None  # always filled
            assert sprint.total_done is None
        else:
            assert _close(sprint.total_done, excel_total_done)

        assert _close(sprint.added, ws[f"G{row}"].value)
        assert _close(sprint.total_added, ws[f"H{row}"].value)

        for col, attr in (("I", "upper"), ("J", "lower"), ("K", "remaining")):
            excel_val = ws[f"{col}{row}"].value
            py_val = getattr(sprint, attr)
            if excel_val in ("#N/A", None) and sprint.sprint_number >= 1:
                # row 12 has real values; later NA after report date
                if excel_val == "#N/A":
                    assert py_val is None
                continue
            if excel_val == "#N/A":
                assert py_val is None
            else:
                assert _close(py_val, excel_val), f"{attr} row {row}: {py_val} != {excel_val}"

        assert sprint.n == int(ws[f"N{row}"].value)

        excel_o = ws[f"O{row}"].value
        if excel_o in ("", None):
            assert sprint.prob_finish_by is None
        else:
            assert _close(sprint.prob_finish_by, excel_o, rel=1e-5, abs_tol=1e-9), (
                f"O{row}: {sprint.prob_finish_by} != {excel_o}"
            )

        excel_m = ws[f"M{row}"].value
        if excel_m in ("", None):
            assert sprint.prob_this_iteration is None
        else:
            assert _close(
                sprint.prob_this_iteration, excel_m, rel=1e-5, abs_tol=1e-9
            ), f"M{row}: {sprint.prob_this_iteration} != {excel_m}"


def test_trend_lines_match_excel(forecast, excel_formulas):
    ws = excel_formulas
    # F12:F26 and L12:L26 are TREND spill; F27 is #N/A in the sheet
    for sprint in forecast.sprints:
        row = 12 + sprint.sprint_number
        excel_f = ws[f"F{row}"].value
        excel_l = ws[f"L{row}"].value

        if excel_f in ("#N/A", None):
            assert sprint.trend_remaining is None
        else:
            assert _close(
                sprint.trend_remaining, excel_f, abs_tol=1e-6
            ), f"F{row}: {sprint.trend_remaining} != {excel_f}"

        if excel_l in ("#N/A", None):
            # L27 empty in sheet
            if sprint.sprint_number == 15:
                assert sprint.trend_added is None
        else:
            assert _close(
                sprint.trend_added, excel_l, abs_tol=1e-6
            ), f"L{row}: {sprint.trend_added} != {excel_l}"


def test_plan_loaded():
    items = load_plan_from_excel(EXCEL)
    assert len(items) == 24
    assert sum(i.estimate for i in items) == 65
    assert items[0].name.startswith("сделать шпаргалку")

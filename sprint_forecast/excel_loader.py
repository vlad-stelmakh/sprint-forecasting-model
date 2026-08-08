"""Load WorkItems from the Plan sheet of sprint_planning.xlsx."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .model import WorkItem


def _cell_date(value) -> date | datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    return None


def load_plan_from_excel(path: str | Path) -> list[WorkItem]:
    """
    Read Plan sheet columns:
    A Название работы, B Оценка, C Дата создания, D Дата завершения
    """
    path = Path(path)
    # data_only=True to get cached DATE() results
    wb = load_workbook(path, data_only=True)
    if "Plan" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Plan' not found in {path}")
    ws = wb["Plan"]

    items: list[WorkItem] = []
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        name, estimate, created, completed = row
        if name is None and estimate is None:
            continue
        if estimate is None or created is None:
            continue
        created_d = _cell_date(created)
        if created_d is None:
            continue
        items.append(
            WorkItem(
                name=str(name),
                estimate=float(estimate),
                created=created_d,
                completed=_cell_date(completed),
            )
        )
    return items


def load_excel_params(path: str | Path) -> dict:
    """Read sprint length and first sprint start from Formulas sheet."""
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb["Formulas"]
    sprint_length = int(ws["B3"].value)
    first_sprint_start = ws["B13"].value
    if isinstance(first_sprint_start, datetime):
        first_sprint_start = first_sprint_start.date()
    return {
        "sprint_length_days": sprint_length,
        "first_sprint_start": first_sprint_start,
    }
